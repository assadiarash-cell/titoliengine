"""API export: CSV generico, Excel formattato, struttura gestionali."""
import csv
import io
import uuid
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from app.api.deps import DbSession
from app.services import journal_service, report_service

router = APIRouter(prefix="/export", tags=["export"])


def _decimal_str(val) -> str:
    """Converte Decimal/None in stringa per CSV/Excel."""
    if val is None:
        return ""
    if isinstance(val, Decimal):
        return str(val.quantize(Decimal("0.01")))
    return str(val)


# ── Export CSV generico ──────────────────────────────────────

@router.get("/journal/csv")
async def export_journal_csv(
    client_id: uuid.UUID,
    session: DbSession,
    fiscal_year: int | None = None,
    status: str | None = None,
):
    """Export scritture contabili in formato CSV.

    Colonne: entry_id, date, description, type, status, line_number,
    account_code, account_name, debit, credit
    """
    entries = await journal_service.list_entries(
        session,
        client_id=client_id,
        fiscal_year=fiscal_year,
        status=status,
    )

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "entry_id", "entry_date", "description", "entry_type", "status",
        "line_number", "account_code", "account_name", "debit", "credit",
    ])

    for entry in entries:
        for line in entry.lines:
            writer.writerow([
                str(entry.id),
                str(entry.entry_date),
                entry.description,
                entry.entry_type,
                entry.status,
                line.line_number,
                line.account_code,
                line.account_name,
                _decimal_str(line.debit),
                _decimal_str(line.credit),
            ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=journal_{client_id}.csv"},
    )


@router.get("/portfolio/csv")
async def export_portfolio_csv(
    client_id: uuid.UUID,
    session: DbSession,
    report_date: date | None = None,
):
    """Export portafoglio in formato CSV."""
    if report_date is None:
        report_date = date.today()

    data = await report_service.portfolio_report(
        session, client_id=client_id, report_date=report_date,
    )

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "isin", "name", "security_type", "classification",
        "quantity", "book_value", "book_value_per_unit",
        "market_price", "market_value", "unrealized_gain_loss",
    ])

    for pos in data["positions"]:
        writer.writerow([
            pos["isin"],
            pos["name"],
            pos["security_type"],
            pos["classification"],
            _decimal_str(pos["quantity"]),
            _decimal_str(pos["book_value"]),
            _decimal_str(pos["book_value_per_unit"]),
            _decimal_str(pos.get("market_price")),
            _decimal_str(pos.get("market_value")),
            _decimal_str(pos.get("unrealized_gain_loss")),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=portfolio_{client_id}.csv"},
    )


# ── Export Excel ─────────────────────────────────────────────

@router.get("/journal/excel")
async def export_journal_excel(
    client_id: uuid.UUID,
    session: DbSession,
    fiscal_year: int | None = None,
    status: str | None = None,
):
    """Export scritture contabili in formato Excel (.xlsx).

    Richiede openpyxl installato.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, numbers
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl non installato. Installa con: pip install openpyxl",
        )

    entries = await journal_service.list_entries(
        session,
        client_id=client_id,
        fiscal_year=fiscal_year,
        status=status,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Scritture Contabili"

    # Header
    headers = [
        "ID Scrittura", "Data", "Descrizione", "Tipo", "Stato",
        "N. Riga", "Codice Conto", "Nome Conto", "Dare", "Avere",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Dati
    row = 2
    for entry in entries:
        for line in entry.lines:
            ws.cell(row=row, column=1, value=str(entry.id))
            ws.cell(row=row, column=2, value=str(entry.entry_date))
            ws.cell(row=row, column=3, value=entry.description)
            ws.cell(row=row, column=4, value=entry.entry_type)
            ws.cell(row=row, column=5, value=entry.status)
            ws.cell(row=row, column=6, value=line.line_number)
            ws.cell(row=row, column=7, value=line.account_code)
            ws.cell(row=row, column=8, value=line.account_name)
            cell_d = ws.cell(row=row, column=9, value=float(line.debit))
            cell_d.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            cell_c = ws.cell(row=row, column=10, value=float(line.credit))
            cell_c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            row += 1

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=journal_{client_id}.xlsx"},
    )


# ── Export per gestionali (PROFIS, TeamSystem) ───────────────

@router.get("/gestionale/profis")
async def export_profis(
    client_id: uuid.UUID,
    session: DbSession,
    fiscal_year: int | None = None,
):
    """Export in formato PROFIS (Sistemi).

    Struttura CSV con campi:
    TIPO_REG;DATA_REG;CAUSALE;COD_CONTO;DESCRIZIONE;IMPORTO_DARE;IMPORTO_AVERE
    """
    entries = await journal_service.list_entries(
        session,
        client_id=client_id,
        fiscal_year=fiscal_year,
        status="posted",
    )

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "TIPO_REG", "DATA_REG", "CAUSALE", "COD_CONTO",
        "DESCRIZIONE", "IMPORTO_DARE", "IMPORTO_AVERE",
    ])

    for entry in entries:
        for line in entry.lines:
            writer.writerow([
                "CN",  # Contabilità Normale
                entry.entry_date.strftime("%d/%m/%Y"),
                entry.entry_type,
                line.account_code,
                line.description or entry.description,
                _decimal_str(line.debit) if line.debit > 0 else "",
                _decimal_str(line.credit) if line.credit > 0 else "",
            ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=profis_{client_id}.csv"},
    )


@router.get("/gestionale/teamsystem")
async def export_teamsystem(
    client_id: uuid.UUID,
    session: DbSession,
    fiscal_year: int | None = None,
):
    """Export in formato TeamSystem.

    Struttura CSV con campi:
    DATA;NUMERO_REG;CAUSALE;SOTTOCONTO;DARE;AVERE;DESCRIZIONE
    """
    entries = await journal_service.list_entries(
        session,
        client_id=client_id,
        fiscal_year=fiscal_year,
        status="posted",
    )

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "DATA", "NUMERO_REG", "CAUSALE", "SOTTOCONTO",
        "DARE", "AVERE", "DESCRIZIONE",
    ])

    for reg_num, entry in enumerate(entries, start=1):
        for line in entry.lines:
            writer.writerow([
                entry.entry_date.strftime("%d/%m/%Y"),
                reg_num,
                entry.entry_type,
                line.account_code,
                _decimal_str(line.debit) if line.debit > 0 else "",
                _decimal_str(line.credit) if line.credit > 0 else "",
                line.description or entry.description,
            ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=teamsystem_{client_id}.csv"},
    )
